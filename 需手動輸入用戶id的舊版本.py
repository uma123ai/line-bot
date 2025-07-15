from flask import Flask, request
import openpyxl
import requests
import threading
import time
import os

app = Flask(__name__)

<<<<<<< HEAD
CHANNEL_ACCESS_TOKEN = ""
USER_ID = ""
=======
CHANNEL_ACCESS_TOKEN = "QOUNRqg0GxzTYA50LypSTmPZenRi43upy0JjEdvU8fXqzgTSsV4AYzwZzpjNaD+akDNyLxPhaJDTmmctgx0d1XbSZfRNc4UpYV3rjI1FvRYrgXf9jwksBC2K4pmmiWva9Zv/qrXyvWqvQ+Ch+J5GDQdB04t89/1O/w1cDnyilFU="
USER_ID = "Ub7dd5c4f68caf6bc36942da58c033491"
>>>>>>> 309933d (新增整合測試版.py)

headers = {
    "Content-Type": "application/json",
    "Authorization": f"Bearer {CHANNEL_ACCESS_TOKEN}"
}

def reply_message(reply_token, text):
    url = "https://api.line.me/v2/bot/message/reply"
    data = {
        "replyToken": reply_token,
        "messages": [{"type": "text", "text": text}]
    }
    r = requests.post(url, headers=headers, json=data)
    print("回覆狀態：", r.status_code)

def push_message(text):
    url = "https://api.line.me/v2/bot/message/push"
    data = {
        "to": USER_ID,
        "messages": [{"type": "text", "text": text}]
    }
    r = requests.post(url, headers=headers, json=data)
    print("推播狀態：", r.status_code)

EXCEL_PATH = os.path.join(os.getcwd(), "出入紀錄.xlsx")

def get_latest_record(filename=EXCEL_PATH):
    try:
        wb = openpyxl.load_workbook(filename)
        sheet = wb.active
        last_row = sheet.max_row
        row_data = [cell.value for cell in sheet[last_row]]
        print("讀到的資料列：", row_data)
        if not row_data or len(row_data) < 4 or any(v is None for v in row_data[:4]):
            return None
        date, time_str, name, status = row_data[:4]
        return f"📋 最新出入紀錄\n🗓 {date} {time_str}\n👤 {name}\n📌 狀態：{status}"
    except Exception as e:
        print(f"讀取 Excel 發生錯誤：{e}")
        return None

def monitor_excel(interval=10):
    last_row_num = 0
    while True:
        try:
            if not os.path.exists(EXCEL_PATH):
                print(f"Excel 檔案不存在：{EXCEL_PATH}")
                time.sleep(interval)
                continue
            wb = openpyxl.load_workbook(EXCEL_PATH)
            sheet = wb.active
            current_last_row = sheet.max_row
            if current_last_row > last_row_num:
                last_row_num = current_last_row
                msg = get_latest_record()
                if msg:
                    print("偵測到新資料，推播中：", msg)
                    push_message(msg)
        except Exception as e:
            print("監控 Excel 發生錯誤，但繼續運行：", e)
        time.sleep(interval)

@app.route("/webhook", methods=["POST"])
def webhook():
    body = request.json
    print("收到 webhook：", body)

    events = body.get("events", [])
    for event in events:
        if event["type"] == "message" and event["message"]["type"] == "text":
            reply_token = event["replyToken"]
            user_id = event["source"]["userId"]
            text = event["message"]["text"]
            print(f"用戶的 userId：{user_id}")
            print(f"訊息內容是：「{text}」")
            reply_message(reply_token, f"收到你的訊息：{text}")

    return "OK"

if __name__ == "__main__":
    # 啟動監控執行緒（daemon模式，主程式結束它會自動結束）
    t = threading.Thread(target=monitor_excel, daemon=True)
    t.start()

    app.run(port=5000)