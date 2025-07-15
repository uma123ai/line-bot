from flask import Flask, request
import requests
from openpyxl import load_workbook
import os

app = Flask(__name__)

# ========== 設定區 ==========
EXCEL_FILE = "出入紀錄.xlsx"
USER_ID_FILE = "user_ids.txt"
CHANNEL_ACCESS_TOKEN = ""

HEADERS = {
    "Content-Type": "application/json",
    "Authorization": f"Bearer {CHANNEL_ACCESS_TOKEN}"
}

# ========== 安全列印 ==========
def safe_print(msg):
    try:
        print(msg)
    except UnicodeEncodeError:
        print(msg.encode("utf-8", errors="ignore").decode("utf-8", errors="ignore"))

# ========== 推播訊息 ==========
def push_message(text, user_id):
    data = {
        "to": user_id,
        "messages": [{"type": "text", "text": str(text)}]
    }
    url = "https://api.line.me/v2/bot/message/push"
    r = requests.post(url, headers=HEADERS, json=data)
    safe_print(f"🔔 推播給 {user_id}：{r.status_code}")

def push_to_all_users(text):
    user_ids = load_user_ids()
    for uid in user_ids:
        push_message(text, uid)

# ========== 儲存/讀取使用者 ID ==========
def save_user_id(user_id):
    if not os.path.exists(USER_ID_FILE):
        with open(USER_ID_FILE, "w"): pass
    with open(USER_ID_FILE, "r+") as f:
        ids = [line.strip() for line in f]
        if user_id not in ids:
            f.write(user_id + "\n")
            safe_print(f"✅ 儲存新 userId：{user_id}")

def load_user_ids():
    if not os.path.exists(USER_ID_FILE):
        return []
    with open(USER_ID_FILE, "r") as f:
        return [line.strip() for line in f]

# ========== 回覆訊息 ==========
def reply_message(reply_token, text):
    url = "https://api.line.me/v2/bot/message/reply"
    data = {
        "replyToken": reply_token,
        "messages": [{"type": "text", "text": str(text)}]
    }
    requests.post(url, headers=HEADERS, json=data)

# ========== 快速選單 ==========
def reply_quick_menu(reply_token):
    url = "https://api.line.me/v2/bot/message/reply"
    data = {
        "replyToken": reply_token,
        "messages": [{
            "type": "text",
            "text": "📋 請選擇操作項目：",
            "quickReply": {
                "items": [
                    {
                        "type": "action",
                        "action": {
                            "type": "message",
                            "label": "查詢最近紀錄",
                            "text": "查詢最近紀錄"
                        }
                    }
                ]
            }
        }]
    }
    requests.post(url, headers=HEADERS, json=data)

# ========== 讀取最近出入紀錄 ==========
def get_latest_entries(n=5):
    try:
        wb = load_workbook(EXCEL_FILE)
        sheet = wb.active
        max_row = sheet.max_row
        messages = []
        for row in range(max(max_row - n + 1, 2), max_row + 1):
            values = [cell.value for cell in sheet[row][:4]]
            if len(values) < 4 or any(v is None for v in values):
                continue
            date, time_str, name, status = values
            msg = f"🕒 {date} {time_str}\n👤 {name}\n📌 狀態：{status}"
            messages.append(msg)
        return "📋 最近出入紀錄：\n\n" + "\n\n".join(messages) if messages else "⚠️ 找不到有效資料。"
    except Exception as e:
        return f"⚠️ 讀取 Excel 發生錯誤：{e}"

# ========== webhook 入口 ==========
@app.route("/webhook", methods=["POST"])
def webhook():
    body = request.json
    events = body.get("events", [])
    for event in events:
        if event["type"] == "message":
            user_id = event["source"]["userId"]
            text = event["message"]["text"]
            reply_token = event["replyToken"]

            save_user_id(user_id)

            if text == "選單":
                reply_quick_menu(reply_token)
            elif text == "查詢最近紀錄":
                msg = get_latest_entries(5)
                reply_message(reply_token, msg)
            else:
                reply_message(reply_token, f"👋 收到訊息：「{text}」\n請輸入「選單」查看操作項目。")
    return "OK"

if __name__ == "__main__":
    app.run(port=5000)
