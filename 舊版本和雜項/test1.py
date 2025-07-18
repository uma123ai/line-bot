# -*- coding: utf-8 -*-
import os
import time
import threading
import requests
from flask import Flask, request
from openpyxl import load_workbook
from pyngrok import ngrok
import cloudinary
from cloudinary.uploader import upload as cloudinary_upload, destroy as cloudinary_destroy

# ========== 設定區 ==========
EXCEL_FILE = "出入紀錄.xlsx"
USER_ID_FILE = "user_ids.txt"
STATE_FILE = "last_row_num.txt"
USER_IMAGE_FILE = "user_images.txt"       # 存 user_id,img_url,public_id
USER_NAME_BIND_FILE = "user_name_bind.txt"  # 存 name,user_id 綁定 Excel姓名與user_id

CHANNEL_ACCESS_TOKEN = "QOUNRqg0GxzTYA50LypSTmPZenRi43upy0JjEdvU8fXqzgTSsV4AYzwZzpjNaD+akDNyLxPhaJDTmmctgx0d1XbSZfRNc4UpYV3rjI1FvRYrgXf9jwksBC2K4pmmiWva9Zv/qrXyvWqvQ+Ch+J5GDQdB04t89/1O/w1cDnyilFU="
CLOUDINARY_CLOUD_NAME = "dijfxhzqd"
CLOUDINARY_API_KEY = "158261511847545"
CLOUDINARY_API_SECRET = "cKny5zb9l2KhpoxIgUfKDOaTxlo"

cloudinary.config(
    cloud_name=CLOUDINARY_CLOUD_NAME,
    api_key=CLOUDINARY_API_KEY,
    api_secret=CLOUDINARY_API_SECRET,
    secure=True
)

HEADERS = {
    "Content-Type": "application/json",
    "Authorization": f"Bearer {CHANNEL_ACCESS_TOKEN}"
}

app = Flask(__name__)
last_row_num = 1  # Excel 讀取起始行


# ----- user_id 存取 -----
def save_user_id(user_id):
    if not os.path.exists(USER_ID_FILE):
        with open(USER_ID_FILE, "w", encoding="utf-8") as f:
            pass
    with open(USER_ID_FILE, "r+", encoding="utf-8") as f:
        ids = [line.strip() for line in f]
        if user_id not in ids:
            f.write(user_id + "\n")
            print(f"✅ 儲存新 userId：{user_id}")


def load_user_ids():
    if not os.path.exists(USER_ID_FILE):
        return []
    with open(USER_ID_FILE, "r", encoding="utf-8") as f:
        return [line.strip() for line in f if line.strip()]


# ----- user_name ↔ user_id 綁定 -----
def save_user_name_bind(name, user_id):
    data = {}
    if os.path.exists(USER_NAME_BIND_FILE):
        with open(USER_NAME_BIND_FILE, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                n, uid = line.split(",", 1)
                data[n] = uid
    # 更新或新增
    data[name] = user_id
    with open(USER_NAME_BIND_FILE, "w", encoding="utf-8") as f:
        for n, uid in data.items():
            f.write(f"{n},{uid}\n")
    print(f"✅ 綁定姓名 {name} -> user_id {user_id}")


def load_user_id_by_name(name):
    if not os.path.exists(USER_NAME_BIND_FILE):
        return None
    with open(USER_NAME_BIND_FILE, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            n, uid = line.split(",", 1)
            if n == name:
                return uid
    return None


# ----- user 圖片存取 -----
def save_user_image(user_id, img_url, public_id):
    data = {}
    if os.path.exists(USER_IMAGE_FILE):
        with open(USER_IMAGE_FILE, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                uid, url, pid = line.split(",", 2)
                data[uid] = (url, pid)
    # 刪除舊圖
    if user_id in data:
        old_public_id = data[user_id][1]
        try:
            cloudinary_destroy(old_public_id)
            print(f"✅ 刪除舊圖片 public_id={old_public_id}")
        except Exception as e:
            print(f"❌ 刪除舊圖片失敗: {e}")

    data[user_id] = (img_url, public_id)
    with open(USER_IMAGE_FILE, "w", encoding="utf-8") as f:
        for uid, (url, pid) in data.items():
            f.write(f"{uid},{url},{pid}\n")
    print(f"✅ 儲存 user({user_id}) 圖片資訊")


def load_user_image(user_id):
    if not os.path.exists(USER_IMAGE_FILE):
        return None, None
    with open(USER_IMAGE_FILE, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            uid, url, pid = line.split(",", 2)
            if uid == user_id:
                return url, pid
    return None, None


# ----- 推播函式 -----
def push_message_text(text, user_id):
    data = {"to": user_id, "messages": [{"type": "text", "text": text}]}
    url = "https://api.line.me/v2/bot/message/push"
    try:
        r = requests.post(url, headers=HEADERS, json=data)
        print(f"🔔 推播給 {user_id}：{r.status_code}")
        if r.status_code == 403:
            print(f"❌ 用戶 {user_id} 已封鎖，跳過推播")
            return False
        return True
    except Exception as e:
        print(f"❌ 推播失敗 {user_id}: {e}")
        return False


def push_message_image(text, image_url, user_id):
    data = {
        "to": user_id,
        "messages": [
            {"type": "text", "text": text},
            {"type": "image", "originalContentUrl": image_url, "previewImageUrl": image_url},
        ],
    }
    url = "https://api.line.me/v2/bot/message/push"
    try:
        r = requests.post(url, headers=HEADERS, json=data)
        print(f"🔔 推播圖片給 {user_id}：{r.status_code}")
        if r.status_code == 403:
            print(f"❌ 用戶 {user_id} 已封鎖，跳過推播")
            return False
        return True
    except Exception as e:
        print(f"❌ 推播失敗 {user_id}: {e}")
        return False


# 推播新增 Excel 資料：每筆資料 -> 所有人都收到文字，綁定姓名者收到圖片
def push_new_record_to_all(date, time_str, name, status):
    text = f"📋 新出入紀錄\n🗓 {date} {time_str}\n👤 {name}\n📌 狀態：{status}"
    user_ids = load_user_ids()

    # 該姓名是否有綁定 user_id
    bind_user_id = load_user_id_by_name(name)

    for uid in user_ids:
        if uid == bind_user_id:
            # 該用戶設定的圖片
            img_url, _ = load_user_image(uid)
            if img_url:
                push_message_image(text, img_url, uid)
            else:
                push_message_text(text, uid)
        else:
            # 其他用戶只收純文字
            push_message_text(text, uid)


# ----- 回覆函式 -----
def reply_message(reply_token, text):
    url = "https://api.line.me/v2/bot/message/reply"
    data = {"replyToken": reply_token, "messages": [{"type": "text", "text": text}]}
    requests.post(url, headers=HEADERS, json=data)


def reply_quick_menu(reply_token):
    url = "https://api.line.me/v2/bot/message/reply"
    data = {
        "replyToken": reply_token,
        "messages": [
            {
                "type": "text",
                "text": "📋 請選擇操作項目：",
                "quickReply": {
                    "items": [
                        {
                            "type": "action",
                            "action": {
                                "type": "message",
                                "label": "查詢最近紀錄",
                                "text": "查詢最近紀錄",
                            },
                        },
                        {
                            "type": "action",
                            "action": {
                                "type": "message",
                                "label": "設定我的圖片",
                                "text": "我要設定圖片",
                            },
                        },
                        {
                            "type": "action",
                            "action": {
                                "type": "message",
                                "label": "綁定姓名",
                                "text": "我要綁定姓名",
                            },
                        },
                    ]
                },
            }
        ],
    }
    requests.post(url, headers=HEADERS, json=data)


# 讀取最近 n 筆出入紀錄
def get_latest_entries(n=5):
    try:
        wb = load_workbook(EXCEL_FILE)
        sheet = wb.active
        max_row = sheet.max_row
        messages = []
        for row in range(max(max_row - n + 1, 2), max_row + 1):
            values = [cell.value for cell in sheet[row][:4]]
            if len(values) < 4 or any(v is None for v in values):
                continue
            date, time_str, name, status = values
            msg = f"🕒 {date} {time_str}\n👤 {name}\n📌 狀態：{status}"
            messages.append(msg)
        return "📋 最近出入紀錄：\n\n" + "\n\n".join(messages) if messages else "⚠️ 找不到有效資料。"
    except Exception as e:
        return f"⚠️ 讀取 Excel 發生錯誤：{e}"


# ----- Webhook 入口 -----
@app.route("/webhook", methods=["POST"])
def webhook():
    body = request.json
    events = body.get("events", [])
    for event in events:
        if event["type"] == "message":
            user_id = event["source"]["userId"]
            reply_token = event["replyToken"]
            save_user_id(user_id)

            msg_type = event["message"]["type"]

            if msg_type == "text":
                text = event["message"]["text"]
                if text == "選單":
                    reply_quick_menu(reply_token)
                elif text == "查詢最近紀錄":
                    msg = get_latest_entries(5)
                    reply_message(reply_token, msg)
                elif text == "我要設定圖片":
                    reply_message(reply_token, "📤 請傳送圖片給我，我會幫你設定為個人圖片。")
                elif text == "我要綁定姓名":
                    reply_message(reply_token, "✍️ 請輸入您想綁定的姓名（必須與 Excel 中姓名欄一致）。")
                else:
                    # 綁定姓名流程：用戶輸入姓名即視為綁定請求（簡易判斷）
                    if text and text != "選單" and text != "查詢最近紀錄" and text != "我要設定圖片" and text != "我要綁定姓名":
                        save_user_name_bind(text, user_id)
                        reply_message(reply_token, f"✅ 您已成功綁定姓名：{text}")
                    else:
                        reply_message(reply_token, f"👋 收到訊息：「{text}」\n請輸入「選單」查看操作項目。")

            elif msg_type == "image":
                message_id = event["message"]["id"]
                image_url = f"https://api-data.line.me/v2/bot/message/{message_id}/content"
                headers = {"Authorization": f"Bearer {CHANNEL_ACCESS_TOKEN}"}
                res = requests.get(image_url, headers=headers)
                if res.status_code == 200:
                    image_data = res.content
                    try:
                        upload_result = cloudinary_upload(image_data)
                        img_url = upload_result["secure_url"]
                        public_id = upload_result["public_id"]

                        save_user_image(user_id, img_url, public_id)

                        reply_message(reply_token, "✅ 圖片已成功設定為您的個人圖片！")
                    except Exception as e:
                        reply_message(reply_token, f"❌ 圖片上傳失敗：{e}")
                else:
                    reply_message(reply_token, f"❌ 下載圖片失敗，狀態碼：{res.status_code}")

    return "OK"


# ----- 監控 Excel 並推播 -----
def monitor_excel(interval=10):
    global last_row_num
    last_row_num = 1
    if os.path.exists(STATE_FILE):
        with open(STATE_FILE, "r", encoding="utf-8") as f:
            try:
                last_row_num = int(f.read())
            except:
                last_row_num = 1

    while True:
        try:
            if not os.path.exists(EXCEL_FILE):
                print(f"❌ 找不到 {EXCEL_FILE}")
                time.sleep(interval)
                continue

            wb = load_workbook(EXCEL_FILE)
            sheet = wb.active
            current_last_row = sheet.max_row

            if current_last_row > last_row_num:
                print(f"🆕 偵測到新增資料：{current_last_row - last_row_num} 列")
                for row in range(last_row_num + 1, current_last_row + 1):
                    values = [cell.value for cell in sheet[row][:4]]
                    if len(values) < 4 or any(v is None for v in values):
                        continue
                    date, time_str, name, status = values
                    push_new_record_to_all(date, time_str, name, status)

                last_row_num = current_last_row
                with open(STATE_FILE, "w", encoding="utf-8") as f:
                    f.write(str(last_row_num))
        except Exception as e:
            print(f"❌ 監控錯誤：{e}")
        time.sleep(interval)


if __name__ == "__main__":
    port = 5000
    public_url = ngrok.connect(port)
    print(f"🌐 Ngrok 公開網址：{public_url}")
    print(f"📡 請設定 LINE Webhook 為：{public_url}/webhook")

    threading.Thread(target=monitor_excel, daemon=True).start()
    app.run(port=port, threaded=True)
