# -*- coding: utf-8 -*-
import os
import time
import threading
import requests
from flask import Flask, request
from openpyxl import load_workbook
from pyngrok import ngrok
import cloudinary
from cloudinary.uploader import upload as cloudinary_upload, destroy as cloudinary_destroy

# ========== 設定區 ==========
EXCEL_FILE = "出入紀錄.xlsx"
USER_ID_FILE = "user_ids.txt"
STATE_FILE = "last_row_num.txt"
USER_SETTINGS_FILE = "user_settings.txt"  # user_id,bind_name,img_url,public_id,mode

CHANNEL_ACCESS_TOKEN = "QOUNRqg0GxzTYA50LypSTmPZenRi43upy0JjEdvU8fXqzgTSsV4AYzwZzpjNaD+akDNyLxPhaJDTmmctgx0d1XbSZfRNc4UpYV3rjI1FvRYrgXf9jwksBC2K4pmmiWva9Zv/qrXyvWqvQ+Ch+J5GDQdB04t89/1O/w1cDnyilFU="
CLOUDINARY_CLOUD_NAME = "dijfxhzqd"
CLOUDINARY_API_KEY = "158261511847545"
CLOUDINARY_API_SECRET = "cKny5zb9l2KhpoxIgUfKDOaTxlo"

# 初始化 Cloudinary 設定
cloudinary.config(
    cloud_name=CLOUDINARY_CLOUD_NAME,
    api_key=CLOUDINARY_API_KEY,
    api_secret=CLOUDINARY_API_SECRET,
    secure=True
)

HEADERS = {
    "Content-Type": "application/json",
    "Authorization": f"Bearer {CHANNEL_ACCESS_TOKEN}"
}

app = Flask(__name__)
last_row_num = 1  # 預設從第 2 行開始

# --------------------------------
# 工具函式區

def save_user_id(user_id):
    if not os.path.exists(USER_ID_FILE):
        with open(USER_ID_FILE, "w", encoding="utf-8") as f:
            pass
    with open(USER_ID_FILE, "r+", encoding="utf-8") as f:
        ids = [line.strip() for line in f]
        if user_id not in ids:
            f.write(user_id + "\n")
            print("✅ 儲存新 userId：", user_id)

def load_user_ids():
    if not os.path.exists(USER_ID_FILE):
        return []
    with open(USER_ID_FILE, "r", encoding="utf-8") as f:
        return [line.strip() for line in f if line.strip()]

# 儲存用戶設定（包含綁定名字、圖片URL、public_id、模式）
def save_user_settings(user_id, bind_name, img_url, public_id, mode="圖文"):
    data = {}
    if os.path.exists(USER_SETTINGS_FILE):
        with open(USER_SETTINGS_FILE, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                uid, name, url, pid, m = line.split(",", 4)
                data[uid] = (name, url, pid, m)

    # 如果已有舊圖片，要先刪除舊圖片 (如果 public_id 不同)
    if user_id in data:
        old_public_id = data[user_id][2]
        if old_public_id != public_id:
            try:
                cloudinary_destroy(old_public_id)
                print(f"✅ 已刪除舊圖片 public_id={old_public_id}")
            except Exception as e:
                print(f"❌ 刪除舊圖片失敗: {e}")

    data[user_id] = (bind_name, img_url, public_id, mode)

    with open(USER_SETTINGS_FILE, "w", encoding="utf-8") as f:
        for uid, (name, url, pid, m) in data.items():
            f.write(f"{uid},{name},{url},{pid},{m}\n")
    print(f"✅ 儲存 user({user_id}) 設定")

def load_user_settings():
    data = {}
    if os.path.exists(USER_SETTINGS_FILE):
        with open(USER_SETTINGS_FILE, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                uid, name, url, pid, m = line.split(",", 4)
                data[uid] = {
                    "name": name,
                    "img_url": url,
                    "public_id": pid,
                    "mode": m
                }
    return data

def get_user_id_by_name(target_name):
    settings = load_user_settings()
    for uid, info in settings.items():
        if info["name"] == target_name:
            return uid
    return None

def push_message_text(text, user_id):
    data = {
        "to": user_id,
        "messages": [{"type": "text", "text": text}]
    }
    url = "https://api.line.me/v2/bot/message/push"
    r = requests.post(url, headers=HEADERS, json=data)
    print(f"🔔 推播給 {user_id}：{r.status_code}")

def push_message_image(text, image_url, user_id):
    data = {
        "to": user_id,
        "messages": [
            {"type": "text", "text": text},
            {
                "type": "image",
                "originalContentUrl": image_url,
                "previewImageUrl": image_url
            }
        ]
    }
    url = "https://api.line.me/v2/bot/message/push"
    r = requests.post(url, headers=HEADERS, json=data)
    print(f"🔔 推播圖片給 {user_id}：{r.status_code}")

def push_to_all_users(text, img_url=None):
    user_ids = load_user_ids()
    settings = load_user_settings()
    for uid in user_ids:
        user_mode = settings.get(uid, {}).get("mode", "圖文")
        if img_url and user_mode == "圖文":
            push_message_image(text, img_url, uid)
        else:
            push_message_text(text, uid)

def reply_message(reply_token, text):
    url = "https://api.line.me/v2/bot/message/reply"
    data = {
        "replyToken": reply_token,
        "messages": [{"type": "text", "text": text}]
    }
    requests.post(url, headers=HEADERS, json=data)

def reply_quick_menu(reply_token):
    url = "https://api.line.me/v2/bot/message/reply"
    data = {
        "replyToken": reply_token,
        "messages": [{
            "type": "text",
            "text": "📋 請選擇操作項目：",
            "quickReply": {
                "items": [
                    {
                        "type": "action",
                        "action": {
                            "type": "message",
                            "label": "查詢最近紀錄",
                            "text": "查詢最近紀錄"
                        }
                    },
                    {
                        "type": "action",
                        "action": {
                            "type": "message",
                            "label": "設定我的圖片與名字",
                            "text": "我要設定圖片與名字"
                        }
                    },
                    {
                        "type": "action",
                        "action": {
                            "type": "message",
                            "label": "切換純文字/圖片模式",
                            "text": "切換模式"
                        }
                    }
                ]
            }
        }]
    }
    requests.post(url, headers=HEADERS, json=data)

def get_latest_entries_for_all():
    try:
        wb = load_workbook(EXCEL_FILE)
        sheet = wb.active
        max_row = sheet.max_row
        last_records = {}
        # 從下往上讀取，找每個名字最後一筆紀錄
        for row in range(max_row, 1, -1):
            values = [cell.value for cell in sheet[row][:4]]
            if len(values) < 4 or any(v is None for v in values):
                continue
            date, time_str, name, status = values
            if name not in last_records:
                last_records[name] = (date, time_str, status)
        if not last_records:
            return "⚠️ 找不到有效資料。"
        msgs = []
        for name, (date, time_str, status) in last_records.items():
            msgs.append(f"🕒 {date} {time_str}\n👤 {name}\n📌 狀態：{status}")
        return "📋 各成員最後一筆出入紀錄：\n\n" + "\n\n".join(msgs)
    except Exception as e:
        return f"⚠️ 讀取 Excel 發生錯誤：{e}"

# --------------------------------
# Webhook 主流程

@app.route("/webhook", methods=["POST"])
def webhook():
    body = request.json
    events = body.get("events", [])
    for event in events:
        if event["type"] == "message":
            user_id = event["source"]["userId"]
            reply_token = event["replyToken"]
            save_user_id(user_id)

            msg_type = event["message"]["type"]

            if msg_type == "text":
                text = event["message"]["text"].strip()
                if text == "選單":
                    reply_quick_menu(reply_token)
                elif text == "查詢最近紀錄":
                    msg = get_latest_entries_for_all()
                    reply_message(reply_token, msg)
                elif text == "我要設定圖片與名字":
                    reply_message(reply_token, "📤 請先輸入您想綁定的名字（必須跟Excel中名字一致），接著再傳送您要設定的圖片。")
                    # 標記此用戶進入設定流程 (可用簡易暫存，或寫入檔案狀態)
                    user_states[user_id] = {"step": "等待名字"}
                elif text == "切換模式":
                    # 讀取設定切換純文字/圖片模式
                    settings = load_user_settings()
                    user_setting = settings.get(user_id)
                    if not user_setting:
                        reply_message(reply_token, "❌ 尚未設定圖片與名字，請先設定。")
                    else:
                        current_mode = user_setting["mode"]
                        new_mode = "純文字" if current_mode == "圖文" else "圖文"
                        # 更新設定
                        save_user_settings(user_id, user_setting["name"], user_setting["img_url"], user_setting["public_id"], new_mode)
                        reply_message(reply_token, f"✅ 已切換為 {new_mode} 模式。")
                else:
                    # 若用戶正在設定流程中
                    if user_id in user_states:
                        step = user_states[user_id].get("step")
                        if step == "等待名字":
                            bind_name = text
                            user_states[user_id]["bind_name"] = bind_name
                            user_states[user_id]["step"] = "等待圖片"
                            reply_message(reply_token, f"✅ 名字設定為：{bind_name}\n請接著傳送您要設定的圖片。")
                        else:
                            reply_message(reply_token, f"👋 收到訊息：「{text}」\n請輸入「選單」查看操作項目。")
                    else:
                        reply_message(reply_token, f"👋 收到訊息：「{text}」\n請輸入「選單」查看操作項目。")

            elif msg_type == "image":
                # 用戶正在設定圖片階段
                if user_id in user_states and user_states[user_id].get("step") == "等待圖片":
                    message_id = event["message"]["id"]
                    image_url = f"https://api-data.line.me/v2/bot/message/{message_id}/content"
                    headers = {"Authorization": f"Bearer {CHANNEL_ACCESS_TOKEN}"}
                    res = requests.get(image_url, headers=headers)
                    if res.status_code == 200:
                        image_data = res.content
                        try:
                            upload_result = cloudinary_upload(image_data)
                            img_url = upload_result["secure_url"]
                            public_id = upload_result["public_id"]

                            bind_name = user_states[user_id].get("bind_name", "")
                            save_user_settings(user_id, bind_name, img_url, public_id, "圖文")

                            reply_message(reply_token, "✅ 圖片與名字設定完成！您現在可使用「查詢最近紀錄」查看回傳。")

                            user_states.pop(user_id, None)
                        except Exception as e:
                            reply_message(reply_token, f"❌ 圖片上傳失敗：{e}")
                    else:
                        reply_message(reply_token, f"❌ 下載圖片失敗，狀態碼：{res.status_code}")
                else:
                    reply_message(reply_token, "📤 請先使用「我要設定圖片與名字」指令開始設定。")

    return "OK"

# --------------------------------
# 監控 Excel 新增資料並推播

def monitor_excel(interval=10):
    global last_row_num
    last_row_num = 1
    if os.path.exists(STATE_FILE):
        with open(STATE_FILE, "r", encoding="utf-8") as f:
            try:
                last_row_num = int(f.read())
            except:
                last_row_num = 1

    while True:
        try:
            if not os.path.exists(EXCEL_FILE):
                print(f"❌ 找不到 {EXCEL_FILE}")
                time.sleep(interval)
                continue

            wb = load_workbook(EXCEL_FILE)
            sheet = wb.active
            current_last_row = sheet.max_row

            if current_last_row > last_row_num:
                print(f"🆕 偵測到新增資料：{current_last_row - last_row_num} 列")
                for row in range(last_row_num + 1, current_last_row + 1):
                    values = [cell.value for cell in sheet[row][:4]]
                    if len(values) < 4 or any(v is None for v in values):
                        continue
                    date, time_str, name, status = values
                    msg = f"📋 新出入紀錄\n🗓 {date} {time_str}\n👤 {name}\n📌 狀態：{status}"

                    # 用名字找 user_id (設定圖片)
                    user_id_for_img = get_user_id_by_name(name)
                    settings = load_user_settings()
                    if user_id_for_img and user_id_for_img in settings:
                        img_url = settings[user_id_for_img].get("img_url")
                        if img_url:
                            push_to_all_users(msg, img_url)
                        else:
                            push_to_all_users(msg)
                    else:
                        push_to_all_users(msg)

                last_row_num = current_last_row
                with open(STATE_FILE, "w", encoding="utf-8") as f:
                    f.write(str(last_row_num))
        except Exception as e:
            print("❌ 監控錯誤：", e)
        time.sleep(interval)

# --------------------------------
# 簡易用戶狀態記錄 (用於設定流程)

user_states = {}

# --------------------------------
# 主程式入口

if __name__ == "__main__":
    port = 5000
    public_url = ngrok.connect(port)
    print(f"🌐 Ngrok 公開網址：{public_url}")
    print(f"📡 請設定 LINE Webhook 為：{public_url}/webhook")

    threading.Thread(target=monitor_excel, daemon=True).start()
    app.run(port=port, threaded=True)
