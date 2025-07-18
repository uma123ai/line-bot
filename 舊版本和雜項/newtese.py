# -*- coding: utf-8 -*-
import os
import time
import threading
import requests
from flask import Flask, request
from openpyxl import load_workbook
from pyngrok import ngrok
import cloudinary
from cloudinary.uploader import upload as cloudinary_upload

# ========== 設定區 ==========
EXCEL_FILE = "出入紀錄.xlsx"
USER_ID_FILE = "user_ids.txt"
USER_IMAGE_FILE = "user_images.txt"
STATE_FILE = "last_row_num.txt"

CHANNEL_ACCESS_TOKEN = "QOUNRqg0GxzTYA50LypSTmPZenRi43upy0JjEdvU8fXqzgTSsV4AYzwZzpjNaD+akDNyLxPhaJDTmmctgx0d1XbSZfRNc4UpYV3rjI1FvRYrgXf9jwksBC2K4pmmiWva9Zv/qrXyvWqvQ+Ch+J5GDQdB04t89/1O/w1cDnyilFU="
CLOUDINARY_CLOUD_NAME = "dijfxhzqd"
CLOUDINARY_API_KEY = "158261511847545"
CLOUDINARY_API_SECRET = "cKny5zb9l2KhpoxIgUfKDOaTxlo"

# 初始化 Cloudinary
cloudinary.config(
    cloud_name=CLOUDINARY_CLOUD_NAME,
    api_key=CLOUDINARY_API_KEY,
    api_secret=CLOUDINARY_API_SECRET,
    secure=True
)

HEADERS = {
    "Content-Type": "application/json",
    "Authorization": f"Bearer {CHANNEL_ACCESS_TOKEN}"
}

app = Flask(__name__)
last_row_num = 1  # 預設從第 2 行開始

# ========== 使用者資料儲存 ==========

def save_user_id(user_id):
    if not os.path.exists(USER_ID_FILE):
        with open(USER_ID_FILE, "w", encoding="utf-8") as f:
            pass
    with open(USER_ID_FILE, "r+", encoding="utf-8") as f:
        ids = [line.strip() for line in f]
        if user_id not in ids:
            f.write(user_id + "\n")
            print("✅ 儲存新 userId：", user_id)

def load_user_ids():
    if not os.path.exists(USER_ID_FILE):
        return []
    with open(USER_ID_FILE, "r", encoding="utf-8") as f:
        return [line.strip() for line in f if line.strip()]

def save_user_image(user_id, image_url):
    data = {}
    if os.path.exists(USER_IMAGE_FILE):
        with open(USER_IMAGE_FILE, "r", encoding="utf-8") as f:
            for line in f:
                if ',' in line:
                    uid, url = line.strip().split(",", 1)
                    data[uid] = url
    data[user_id] = image_url
    with open(USER_IMAGE_FILE, "w", encoding="utf-8") as f:
        for uid, url in data.items():
            f.write(f"{uid},{url}\n")
    print(f"✅ 儲存 {user_id} 的圖片 URL")

def get_user_image(user_id):
    if os.path.exists(USER_IMAGE_FILE):
        with open(USER_IMAGE_FILE, "r", encoding="utf-8") as f:
            for line in f:
                if ',' in line:
                    uid, url = line.strip().split(",", 1)
                    if uid == user_id:
                        return url
    return None

# ========== LINE 推播/回覆 ==========

def push_message_text(text, user_id):
    data = {
        "to": user_id,
        "messages": [{"type": "text", "text": text}]
    }
    url = "https://api.line.me/v2/bot/message/push"
    r = requests.post(url, headers=HEADERS, json=data)
    print(f"🔔 推播給 {user_id}：{r.status_code}")

def push_message_image(text, image_url, user_id):
    data = {
        "to": user_id,
        "messages": [
            {"type": "text", "text": text},
            {
                "type": "image",
                "originalContentUrl": image_url,
                "previewImageUrl": image_url
            }
        ]
    }
    url = "https://api.line.me/v2/bot/message/push"
    r = requests.post(url, headers=HEADERS, json=data)
    print(f"🔔 推播圖片給 {user_id}：{r.status_code}")

def push_to_all_users(text):
    user_ids = load_user_ids()
    for uid in user_ids:
        img_url = get_user_image(uid)
        if img_url:
            push_message_image(text, img_url, uid)
        else:
            push_message_text(text, uid)

def reply_message(reply_token, text):
    url = "https://api.line.me/v2/bot/message/reply"
    data = {
        "replyToken": reply_token,
        "messages": [{"type": "text", "text": text}]
    }
    requests.post(url, headers=HEADERS, json=data)

def reply_quick_menu(reply_token):
    url = "https://api.line.me/v2/bot/message/reply"
    data = {
        "replyToken": reply_token,
        "messages": [{
            "type": "text",
            "text": "📋 請選擇操作項目：",
            "quickReply": {
                "items": [
                    {
                        "type": "action",
                        "action": {
                            "type": "message",
                            "label": "查詢最近紀錄",
                            "text": "查詢最近紀錄"
                        }
                    }
                ]
            }
        }]
    }
    requests.post(url, headers=HEADERS, json=data)

# ========== 下載並上傳圖片到 Cloudinary ==========

def download_and_upload_image(message_id):
    url = f"https://api-data.line.me/v2/bot/message/{message_id}/content"
    headers = {"Authorization": f"Bearer {CHANNEL_ACCESS_TOKEN}"}
    res = requests.get(url, headers=headers)
    if res.status_code == 200:
        try:
            upload_result = cloudinary_upload(res.content)
            print("✅ 圖片上傳成功，URL:", upload_result["secure_url"])
            return upload_result["secure_url"]
        except Exception as e:
            print("❌ Cloudinary 上傳錯誤：", e)
            return None
    else:
        print(f"❌ 下載圖片失敗，狀態碼：{res.status_code}")
        return None

# ========== 讀取 Excel 最近紀錄 ==========

def get_latest_entries(n=5):
    try:
        wb = load_workbook(EXCEL_FILE)
        sheet = wb.active
        max_row = sheet.max_row
        messages = []
        for row in range(max(max_row - n + 1, 2), max_row + 1):
            values = [cell.value for cell in sheet[row][:4]]
            if len(values) < 4 or any(v is None for v in values):
                continue
            date, time_str, name, status = values
            msg = f"🕒 {date} {time_str}\n👤 {name}\n📌 狀態：{status}"
            messages.append(msg)
        return "📋 最近出入紀錄：\n\n" + "\n\n".join(messages) if messages else "⚠️ 找不到有效資料。"
    except Exception as e:
        return f"⚠️ 讀取 Excel 發生錯誤：{e}"

# ========== Webhook 入口 ==========

@app.route("/webhook", methods=["POST"])
def webhook():
    body = request.json
    events = body.get("events", [])
    for event in events:
        if event["type"] == "message":
            user_id = event["source"]["userId"]
            msg_type = event["message"]["type"]
            reply_token = event["replyToken"]

            save_user_id(user_id)

            if msg_type == "text":
                text = event["message"]["text"]
                if text == "選單":
                    reply_quick_menu(reply_token)
                elif text == "查詢最近紀錄":
                    msg = get_latest_entries(5)
                    reply_message(reply_token, msg)
                else:
                    reply_message(reply_token, f"👋 收到訊息：「{text}」\n請輸入「選單」查看操作項目。")

            elif msg_type == "image":
                message_id = event["message"]["id"]
                img_url = download_and_upload_image(message_id)
                if img_url:
                    save_user_image(user_id, img_url)
                    reply_message(reply_token, "✅ 圖片已上傳並設定成功！")
                else:
                    reply_message(reply_token, "❌ 圖片上傳失敗，請稍後再試。")

    return "OK"

# ========== 監控 Excel 新增資料推播 ==========

def monitor_excel(interval=10):
    global last_row_num
    last_row_num = 1
    if os.path.exists(STATE_FILE):
        with open(STATE_FILE, "r", encoding="utf-8") as f:
            try:
                last_row_num = int(f.read())
            except:
                last_row_num = 1

    while True:
        try:
            if not os.path.exists(EXCEL_FILE):
                print(f"❌ 找不到 {EXCEL_FILE}")
                time.sleep(interval)
                continue

            wb = load_workbook(EXCEL_FILE)
            sheet = wb.active
            current_last_row = sheet.max_row

            if current_last_row > last_row_num:
                print(f"🆕 偵測到新增資料：{current_last_row - last_row_num} 列")
                for row in range(last_row_num + 1, current_last_row + 1):
                    values = [cell.value for cell in sheet[row][:4]]
                    if len(values) < 4 or any(v is None for v in values):
                        continue
                    date, time_str, name, status = values
                    msg = f"📋 新出入紀錄\n🗓 {date} {time_str}\n👤 {name}\n📌 狀態：{status}"
                    push_to_all_users(msg)
                last_row_num = current_last_row
                with open(STATE_FILE, "w", encoding="utf-8") as f:
                    f.write(str(last_row_num))
        except Exception as e:
            print("❌ 監控錯誤：", e)
        time.sleep(interval)

# ========== 主程式 ==========
if __name__ == "__main__":
    port = 5000
    public_url = ngrok.connect(port)
    print(f"🌐 Ngrok 公開網址：{public_url}")
    print(f"📡 請設定 LINE Webhook 為：{public_url}/webhook")

    threading.Thread(target=monitor_excel, daemon=True).start()
    app.run(port=port, threaded=True)
