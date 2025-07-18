# -*- coding: utf-8 -*-
import os
import time
import threading
import requests
from flask import Flask, request
from openpyxl import load_workbook
from pyngrok import ngrok
import cloudinary
from cloudinary.uploader import upload as cloudinary_upload, destroy as cloudinary_destroy

# ========== 設定區 ==========
EXCEL_FILE = "出入紀錄.xlsx"
USER_ID_FILE = "user_ids.txt"
STATE_FILE = "last_row_num.txt"
USER_IMAGE_FILE = "user_images.txt"  # 用來存 user_id,img_url,public_id

CHANNEL_ACCESS_TOKEN = ""
CLOUDINARY_CLOUD_NAME = ""
CLOUDINARY_API_KEY = ""
CLOUDINARY_API_SECRET = ""

# 初始化 Cloudinary 設定
cloudinary.config(
    cloud_name=CLOUDINARY_CLOUD_NAME,
    api_key=CLOUDINARY_API_KEY,
    api_secret=CLOUDINARY_API_SECRET,
    secure=True
)

HEADERS = {
    "Content-Type": "application/json",
    "Authorization": f"Bearer {CHANNEL_ACCESS_TOKEN}"
}

app = Flask(__name__)
last_row_num = 1  # 從第2行開始讀Excel

def clean_user_image_file():
    """啟動時清理 user_images.txt 裡不符合格式的行，避免拆解錯誤"""
    if not os.path.exists(USER_IMAGE_FILE):
        return
    valid_lines = []
    with open(USER_IMAGE_FILE, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line.count(",") == 2:
                valid_lines.append(line)
            else:
                print(f"⚠️ 移除無效圖片紀錄：{line}")
    with open(USER_IMAGE_FILE, "w", encoding="utf-8") as f:
        for line in valid_lines:
            f.write(line + "\n")

# 儲存 userId（避免重複）
def save_user_id(user_id):
    if not os.path.exists(USER_ID_FILE):
        with open(USER_ID_FILE, "w", encoding="utf-8") as f:
            pass
    with open(USER_ID_FILE, "r+", encoding="utf-8") as f:
        ids = [line.strip() for line in f]
        if user_id not in ids:
            f.write(user_id + "\n")
            print(f"✅ 儲存新 userId：{user_id}")

def load_user_ids():
    if not os.path.exists(USER_ID_FILE):
        return []
    with open(USER_ID_FILE, "r", encoding="utf-8") as f:
        return [line.strip() for line in f if line.strip()]

# 儲存用戶圖片資訊，並刪除舊圖
def save_user_image(user_id, img_url, public_id):
    data = {}
    if os.path.exists(USER_IMAGE_FILE):
        with open(USER_IMAGE_FILE, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                uid, url, pid = line.split(",", 2)
                data[uid] = (url, pid)
    # 刪除舊圖片
    if user_id in data:
        old_public_id = data[user_id][1]
        try:
            cloudinary_destroy(old_public_id)
            print(f"✅ 已刪除舊圖片 public_id={old_public_id}")
        except Exception as e:
            print(f"❌ 刪除舊圖片失敗: {e}")

    data[user_id] = (img_url, public_id)
    with open(USER_IMAGE_FILE, "w", encoding="utf-8") as f:
        for uid, (url, pid) in data.items():
            f.write(f"{uid},{url},{pid}\n")
    print(f"✅ 儲存 user({user_id}) 圖片資訊")

def load_user_image(user_id):
    if not os.path.exists(USER_IMAGE_FILE):
        return None, None
    with open(USER_IMAGE_FILE, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            uid, url, pid = line.split(",", 2)
            if uid == user_id:
                return url, pid
    return None, None

def push_message_text(text, user_id):
    data = {
        "to": user_id,
        "messages": [{"type": "text", "text": text}]
    }
    url = "https://api.line.me/v2/bot/message/push"
    try:
        r = requests.post(url, headers=HEADERS, json=data)
        if r.status_code != 200:
            print(f"⚠️ 傳送失敗 user_id={user_id}, 狀態碼={r.status_code}, 原因={r.text}")
    except Exception as e:
        print(f"❌ 發送訊息時出錯：{e}")

def push_message_image(text, image_url, user_id):
    data = {
        "to": user_id,
        "messages": [
            {"type": "text", "text": text},
            {
                "type": "image",
                "originalContentUrl": image_url,
                "previewImageUrl": image_url
            }
        ]
    }
    url = "https://api.line.me/v2/bot/message/push"
    try:
        r = requests.post(url, headers=HEADERS, json=data)
        if r.status_code != 200:
            print(f"⚠️ 傳送圖片失敗 user_id={user_id}, 狀態碼={r.status_code}, 原因={r.text}")
    except Exception as e:
        print(f"❌ 發送圖片時出錯：{e}")

def push_to_all_users(text):
    user_ids = load_user_ids()
    for uid in user_ids:
        img_url, _ = load_user_image(uid)
        if img_url:
            push_message_image(text, img_url, uid)
        else:
            push_message_text(text, uid)

def reply_message(reply_token, text):
    url = "https://api.line.me/v2/bot/message/reply"
    data = {
        "replyToken": reply_token,
        "messages": [{"type": "text", "text": text}]
    }
    try:
        requests.post(url, headers=HEADERS, json=data)
    except Exception as e:
        print(f"❌ 回覆訊息錯誤：{e}")

def reply_quick_menu(reply_token):
    url = "https://api.line.me/v2/bot/message/reply"
    data = {
        "replyToken": reply_token,
        "messages": [{
            "type": "text",
            "text": "📋 請選擇操作項目：",
            "quickReply": {
                "items": [
                    {
                        "type": "action",
                        "action": {
                            "type": "message",
                            "label": "查詢最近紀錄",
                            "text": "查詢最近紀錄"
                        }
                    },
                    {
                        "type": "action",
                        "action": {
                            "type": "message",
                            "label": "設定我的圖片",
                            "text": "我要設定圖片"
                        }
                    }
                ]
            }
        }]
    }
    try:
        requests.post(url, headers=HEADERS, json=data)
    except Exception as e:
        print(f"❌ 回覆選單錯誤：{e}")

def get_latest_entries(n=5):
    try:
        wb = load_workbook(EXCEL_FILE)
        sheet = wb.active
        max_row = sheet.max_row
        messages = []
        for row in range(max(max_row - n + 1, 2), max_row + 1):
            values = [cell.value for cell in sheet[row][:4]]
            if len(values) < 4 or any(v is None for v in values):
                continue
            date, time_str, name, status = values
            msg = f"🕒 {date} {time_str}\n👤 {name}\n📌 狀態：{status}"
            messages.append(msg)
        return "📋 最近出入紀錄：\n\n" + "\n\n".join(messages) if messages else "⚠️ 找不到有效資料。"
    except Exception as e:
        return f"⚠️ 讀取 Excel 發生錯誤：{e}"

@app.route("/webhook", methods=["POST"])
def webhook():
    body = request.json
    events = body.get("events", [])
    for event in events:
        if event["type"] == "message":
            user_id = event["source"]["userId"]
            reply_token = event["replyToken"]
            save_user_id(user_id)

            msg_type = event["message"]["type"]

            if msg_type == "text":
                text = event["message"]["text"]
                if text == "選單":
                    reply_quick_menu(reply_token)
                elif text == "查詢最近紀錄":
                    msg = get_latest_entries(5)
                    reply_message(reply_token, msg)
                elif text == "我要設定圖片":
                    reply_message(reply_token, "📤 請傳送圖片給我，我會幫你設定為個人圖片。")
                else:
                    reply_message(reply_token, f"👋 收到訊息：「{text}」\n請輸入「選單」查看操作項目。")

            elif msg_type == "image":
                # 下載圖片
                message_id = event["message"]["id"]
                image_url = f"https://api-data.line.me/v2/bot/message/{message_id}/content"
                headers = {"Authorization": f"Bearer {CHANNEL_ACCESS_TOKEN}"}
                res = requests.get(image_url, headers=headers)
                if res.status_code == 200:
                    image_data = res.content
                    try:
                        upload_result = cloudinary_upload(image_data)
                        img_url = upload_result["secure_url"]
                        public_id = upload_result["public_id"]

                        save_user_image(user_id, img_url, public_id)

                        reply_message(reply_token, "✅ 圖片已成功設定為您的個人圖片！")
                    except Exception as e:
                        reply_message(reply_token, f"❌ 圖片上傳失敗：{e}")
                else:
                    reply_message(reply_token, f"❌ 下載圖片失敗，狀態碼：{res.status_code}")

    return "OK"

def monitor_excel(interval=10):
    global last_row_num
    last_row_num = 1
    if os.path.exists(STATE_FILE):
        with open(STATE_FILE, "r", encoding="utf-8") as f:
            try:
                last_row_num = int(f.read())
            except:
                last_row_num = 1

    while True:
        try:
            if not os.path.exists(EXCEL_FILE):
                print(f"❌ 找不到 {EXCEL_FILE}")
                time.sleep(interval)
                continue

            wb = load_workbook(EXCEL_FILE)
            sheet = wb.active
            current_last_row = sheet.max_row

            if current_last_row > last_row_num:
                print(f"🆕 偵測到新增資料：{current_last_row - last_row_num} 列")
                for row in range(last_row_num + 1, current_last_row + 1):
                    values = [cell.value for cell in sheet[row][:4]]
                    if len(values) < 4 or any(v is None for v in values):
                        continue
                    date, time_str, name, status = values
                    msg = f"📋 新出入紀錄\n🗓 {date} {time_str}\n👤 {name}\n📌 狀態：{status}"
                    push_to_all_users(msg)
                last_row_num = current_last_row
                with open(STATE_FILE, "w", encoding="utf-8") as f:
                    f.write(str(last_row_num))
        except Exception as e:
            print("❌ 監控錯誤：", e)
        time.sleep(interval)

if __name__ == "__main__":
    clean_user_image_file()
    port = 5000
    public_url = ngrok.connect(port)
    print(f"🌐 Ngrok 公開網址：{public_url}")
    print(f"📡 請設定 LINE Webhook 為：{public_url}/webhook")

    threading.Thread(target=monitor_excel, daemon=True).start()
    app.run(port=port, threaded=True)
