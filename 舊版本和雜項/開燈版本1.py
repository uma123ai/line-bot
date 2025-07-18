# -*- coding: utf-8 -*-
import os
import time
from openpyxl import load_workbook
from linebot import LineBotApi
from linebot.models import TextSendMessage

LIGHT_CONTROL_FILE = "燈控紀錄.xlsx"
CHANNEL_ACCESS_TOKEN = "QOUNRqg0GxzTYA50LypSTmPZenRi43upy0JjEdvU8fXqzgTSsV4AYzwZzpjNaD+akDNyLxPhaJDTmmctgx0d1XbSZfRNc4UpYV3rjI1FvRYrgXf9jwksBC2K4pmmiWva9Zv/qrXyvWqvQ+Ch+J5GDQdB04t89/1O/w1cDnyilFU="

# 初始化 LINE Bot API（用於推播）
line_bot_api = LineBotApi(CHANNEL_ACCESS_TOKEN)

def load_user_ids():
    USER_ID_FILE = "user_ids.txt"
    if not os.path.exists(USER_ID_FILE):
        return []
    with open(USER_ID_FILE, "r", encoding="utf-8") as f:
        return [line.strip() for line in f if line.strip()]

def push_to_all_users(text):
    user_ids = load_user_ids()
    for uid in user_ids:
        try:
            line_bot_api.push_message(uid, TextSendMessage(text=text))
            print(f"🔔 推播給 {uid}：成功")
        except Exception as e:
            print(f"🔔 推播給 {uid} 失敗：{e}")

def check_light_status():
    try:
        if not os.path.exists(LIGHT_CONTROL_FILE):
            print(f"❌ 找不到 {LIGHT_CONTROL_FILE}")
            return False

        wb = load_workbook(LIGHT_CONTROL_FILE)
        sheet = wb.active
        max_row = sheet.max_row

        if max_row <= 1:  # 第一行為標頭
            print("⚠️ 沒有燈控記錄")
            return False

        latest_row = sheet[max_row]
        values = [cell.value for cell in latest_row[:4]]
        if len(values) < 4 or any(v is None for v in values):
            print("⚠️ 最新記錄無效")
            return False

        date, time_str, user_id, action = values
        is_light_on = (action == "開燈")
        status_msg = f"🕒 最新記錄：{date} {time_str} | 用戶ID：{user_id} | 動作：{action}\n💡 燈狀態：{'開' if is_light_on else '關'}"
        print(status_msg)
        
        # 推播燈狀態給所有用戶
        push_to_all_users(status_msg)
        return is_light_on

    except Exception as e:
        print(f"❌ 讀取燈控記錄失敗：{e}")
        return False

if __name__ == "__main__":
    while True:
        is_light_on = check_light_status()
        if is_light_on:
            print("🎉 目標達成：燈已開啟！")
        else:
            print("🔌 燈目前關閉。")
        time.sleep(10)  # 每 10 秒檢查一次