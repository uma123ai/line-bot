# -*- coding: utf-8 -*-
import os
import time
import threading
import requests
from flask import Flask, request
from openpyxl import load_workbook
from pyngrok import ngrok
import cloudinary
from cloudinary.uploader import upload as cloudinary_upload, destroy as cloudinary_destroy

# ========== 設定區 ==========
EXCEL_FILE = "出入紀錄.xlsx"
USER_ID_FILE = "user_ids.txt"
STATE_FILE = "last_row_num.txt"
USER_PROFILE_FILE = "user_profile.txt"

CHANNEL_ACCESS_TOKEN = "QOUNRqg0GxzTYA50LypSTmPZenRi43upy0JjEdvU8fXqzgTSsV4AYzwZzpjNaD+akDNyLxPhaJDTmmctgx0d1XbSZfRNc4UpYV3rjI1FvRYrgXf9jwksBC2K4pmmiWva9Zv/qrXyvWqvQ+Ch+J5GDQdB04t89/1O/w1cDnyilFU="
CLOUDINARY_CLOUD_NAME = "dijfxhzqd"
CLOUDINARY_API_KEY = "158261511847545"
CLOUDINARY_API_SECRET = "cKny5zb9l2KhpoxIgUfKDOaTxlo"

cloudinary.config(
    cloud_name=CLOUDINARY_CLOUD_NAME,
    api_key=CLOUDINARY_API_KEY,
    api_secret=CLOUDINARY_API_SECRET,
    secure=True
)

HEADERS = {
    "Content-Type": "application/json",
    "Authorization": f"Bearer {CHANNEL_ACCESS_TOKEN}"
}

app = Flask(__name__)
last_row_num = 1

def save_user_id(user_id):
    if not os.path.exists(USER_ID_FILE):
        open(USER_ID_FILE, "w", encoding="utf-8").close()
    with open(USER_ID_FILE, "r+", encoding="utf-8") as f:
        ids = [line.strip() for line in f]
        if user_id not in ids:
            f.write(user_id + "\n")

def load_user_ids():
    if not os.path.exists(USER_ID_FILE): return []
    with open(USER_ID_FILE, "r", encoding="utf-8") as f:
        return [line.strip() for line in f if line.strip()]

def save_user_profile(user_id, img_url=None, public_id=None, mode=None):
    data = {}
    if os.path.exists(USER_PROFILE_FILE):
        with open(USER_PROFILE_FILE, "r", encoding="utf-8") as f:
            for line in f:
                parts = line.strip().split(",", 3)
                data[parts[0]] = parts[1:] if len(parts) == 4 else [None, None, "image"]
    if img_url and public_id:
        old = data.get(user_id)
        if old and old[1]:
            try:
                cloudinary_destroy(old[1])
            except: pass
        data[user_id] = [img_url, public_id, mode or "image"]
    elif mode:
        current = data.get(user_id, [None, None, "image"])
        data[user_id] = [current[0], current[1], mode]
    with open(USER_PROFILE_FILE, "w", encoding="utf-8") as f:
        for uid, (url, pid, m) in data.items():
            f.write(f"{uid},{url or ''},{pid or ''},{m}\n")

def load_user_profile(user_id):
    if not os.path.exists(USER_PROFILE_FILE):
        return None, None, "image"
    with open(USER_PROFILE_FILE, "r", encoding="utf-8") as f:
        for line in f:
            parts = line.strip().split(",", 3)
            if parts[0] == user_id:
                return parts[1] or None, parts[2] or None, parts[3] if len(parts) == 4 else "image"
    return None, None, "image"

def push_message_text(text, user_id):
    data = {"to": user_id, "messages": [{"type": "text", "text": text}]}
    r = requests.post("https://api.line.me/v2/bot/message/push", headers=HEADERS, json=data)

def push_message_image(text, image_url, user_id):
    data = {
        "to": user_id,
        "messages": [
            {"type": "text", "text": text},
            {"type": "image", "originalContentUrl": image_url, "previewImageUrl": image_url}
        ]
    }
    r = requests.post("https://api.line.me/v2/bot/message/push", headers=HEADERS, json=data)

def push_to_all_users(text):
    for uid in load_user_ids():
        img_url, _, mode = load_user_profile(uid)
        if mode == "text" or not img_url:
            push_message_text(text, uid)
        else:
            push_message_image(text, img_url, uid)

def reply_message(token, text):
    data = {"replyToken": token, "messages": [{"type": "text", "text": text}]}
    requests.post("https://api.line.me/v2/bot/message/reply", headers=HEADERS, json=data)

def reply_quick_menu(token):
    items = [
        {"type": "action", "action": {"type": "message", "label": "查詢最近紀錄", "text": "查詢最近紀錄"}},
        {"type": "action", "action": {"type": "message", "label": "設定我的圖片", "text": "我要設定圖片"}},
        {"type": "action", "action": {"type": "message", "label": "純文字模式", "text": "純文字模式"}},
        {"type": "action", "action": {"type": "message", "label": "圖片模式", "text": "圖片模式"}},
    ]
    reply = {"replyToken": token, "messages": [{"type": "text", "text": "📋 請選擇操作項目：", "quickReply": {"items": items}}]}
    requests.post("https://api.line.me/v2/bot/message/reply", headers=HEADERS, json=reply)

def get_latest_entries(n=5):
    try:
        wb = load_workbook(EXCEL_FILE)
        sheet = wb.active
        max_row = sheet.max_row
        messages = []
        for row in range(max(max_row - n + 1, 2), max_row + 1):
            values = [cell.value for cell in sheet[row][:4]]
            if None in values or len(values) < 4: continue
            date, time_str, name, status = values
            messages.append(f"🕒 {date} {time_str}\n👤 {name}\n📌 狀態：{status}")
        return "\n\n".join(messages) if messages else "⚠️ 找不到有效資料。"
    except Exception as e:
        return f"⚠️ 讀取 Excel 發生錯誤：{e}"

@app.route("/webhook", methods=["POST"])
def webhook():
    body = request.json
    for event in body.get("events", []):
        if event["type"] == "message":
            user_id = event["source"]["userId"]
            token = event["replyToken"]
            save_user_id(user_id)
            m_type = event["message"]["type"]

            if m_type == "text":
                text = event["message"]["text"]
                if text == "選單": reply_quick_menu(token)
                elif text == "查詢最近紀錄": reply_message(token, get_latest_entries())
                elif text == "我要設定圖片": reply_message(token, "📤 請傳送圖片給我。")
                elif text == "純文字模式":
                    save_user_profile(user_id, mode="text")
                    reply_message(token, "✅ 您已切換為純文字模式。")
                elif text == "圖片模式":
                    save_user_profile(user_id, mode="image")
                    reply_message(token, "✅ 您已切換為圖片模式。")
                else:
                    reply_message(token, f"👋 收到訊息：「{text}」\n請輸入「選單」查看操作項目。")

            elif m_type == "image":
                mid = event["message"]["id"]
                res = requests.get(f"https://api-data.line.me/v2/bot/message/{mid}/content", headers={"Authorization": f"Bearer {CHANNEL_ACCESS_TOKEN}"})
                if res.status_code == 200:
                    try:
                        result = cloudinary_upload(res.content)
                        save_user_profile(user_id, result["secure_url"], result["public_id"])
                        reply_message(token, "✅ 圖片已成功設定！")
                    except Exception as e:
                        reply_message(token, f"❌ 上傳失敗：{e}")
                else:
                    reply_message(token, "❌ 下載圖片失敗。")
    return "OK"

def monitor_excel(interval=10):
    global last_row_num
    if os.path.exists(STATE_FILE):
        try:
            last_row_num = int(open(STATE_FILE).read())
        except:
            last_row_num = 1

    while True:
        try:
            if not os.path.exists(EXCEL_FILE):
                time.sleep(interval)
                continue

            wb = load_workbook(EXCEL_FILE)
            sheet = wb.active
            current_row = sheet.max_row
            if current_row > last_row_num:
                for r in range(last_row_num + 1, current_row + 1):
                    values = [cell.value for cell in sheet[r][:4]]
                    if None in values: continue
                    d, t, n, s = values
                    msg = f"📋 新出入紀錄\n🗓 {d} {t}\n👤 {n}\n📌 狀態：{s}"
                    push_to_all_users(msg)
                last_row_num = current_row
                open(STATE_FILE, "w", encoding="utf-8").write(str(last_row_num))
        except Exception as e:
            print("❌ 錯誤：", e)
        time.sleep(interval)

if __name__ == "__main__":
    port = 5000
    public_url = ngrok.connect(port)
    print(f"🌐 Ngrok 公開網址：{public_url}/webhook")
    threading.Thread(target=monitor_excel, daemon=True).start()
    app.run(port=port, threaded=True)
