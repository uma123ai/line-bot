# -*- coding: utf-8 -*-
import os
import time
import threading
import requests
from flask import Flask, request
from openpyxl import load_workbook
from pyngrok import ngrok

# ========== 設定區 ==========
EXCEL_FILE = "出入紀錄.xlsx"
USER_ID_FILE = "user_ids.txt"
STATE_FILE = "last_row_num.txt"
CHANNEL_ACCESS_TOKEN = "QOUNRqg0GxzTYA50LypSTmPZenRi43upy0JjEdvU8fXqzgTSsV4AYzwZzpjNaD+akDNyLxPhaJDTmmctgx0d1XbSZfRNc4UpYV3rjI1FvRYrgXf9jwksBC2K4pmmiWva9Zv/qrXyvWqvQ+Ch+J5GDQdB04t89/1O/w1cDnyilFU="

HEADERS = {
    "Content-Type": "application/json",
    "Authorization": f"Bearer {CHANNEL_ACCESS_TOKEN}"
}

# 固定圖片網址（請改成你的Cloudinary圖片URL）
IMAGE_URL = "https://res.cloudinary.com/dijfxhzqd/image/upload/v1752544844/samples/smile.jpg"

app = Flask(__name__)
last_row_num = 1  # 預設從第 2 行開始

# 儲存 userId（避免重複）
def save_user_id(user_id):
    if not os.path.exists(USER_ID_FILE):
        with open(USER_ID_FILE, "w", encoding="utf-8") as f:
            pass
    with open(USER_ID_FILE, "r+", encoding="utf-8") as f:
        ids = [line.strip() for line in f]
        if user_id not in ids:
            f.write(user_id + "\n")
            print("✅ 儲存新 userId：", user_id)

def load_user_ids():
    if not os.path.exists(USER_ID_FILE):
        return []
    with open(USER_ID_FILE, "r", encoding="utf-8") as f:
        return [line.strip() for line in f if line.strip()]

# 推播純文字訊息給單一使用者
def push_message_text(text, user_id):
    data = {
        "to": user_id,
        "messages": [{"type": "text", "text": text}]
    }
    url = "https://api.line.me/v2/bot/message/push"
    r = requests.post(url, headers=HEADERS, json=data)
    print(f"🔔 推播給 {user_id}：{r.status_code}")

# 推播文字＋圖片訊息給單一使用者
def push_message_image(text, image_url, user_id):
    data = {
        "to": user_id,
        "messages": [
            {"type": "text", "text": text},
            {
                "type": "image",
                "originalContentUrl": image_url,
                "previewImageUrl": image_url
            }
        ]
    }
    url = "https://api.line.me/v2/bot/message/push"
    r = requests.post(url, headers=HEADERS, json=data)
    print(f"🔔 推播圖片給 {user_id}：{r.status_code}")

# 推播給所有用戶（圖片參數預設 None，不帶圖片就純文字）
def push_to_all_users(text, image_url=None):
    user_ids = load_user_ids()
    for uid in user_ids:
        if image_url:
            push_message_image(text, image_url, uid)
        else:
            push_message_text(text, uid)

# 回覆純文字訊息
def reply_message(reply_token, text):
    url = "https://api.line.me/v2/bot/message/reply"
    data = {
        "replyToken": reply_token,
        "messages": [{"type": "text", "text": text}]
    }
    requests.post(url, headers=HEADERS, json=data)

# 回覆快速選單
def reply_quick_menu(reply_token):
    url = "https://api.line.me/v2/bot/message/reply"
    data = {
        "replyToken": reply_token,
        "messages": [{
            "type": "text",
            "text": "📋 請選擇操作項目：",
            "quickReply": {
                "items": [
                    {
                        "type": "action",
                        "action": {
                            "type": "message",
                            "label": "查詢最近紀錄",
                            "text": "查詢最近紀錄"
                        }
                    }
                ]
            }
        }]
    }
    requests.post(url, headers=HEADERS, json=data)

# 讀取最近 n 筆出入紀錄
def get_latest_entries(n=5):
    try:
        wb = load_workbook(EXCEL_FILE)
        sheet = wb.active
        max_row = sheet.max_row
        messages = []
        for row in range(max(max_row - n + 1, 2), max_row + 1):
            values = [cell.value for cell in sheet[row][:4]]
            if len(values) < 4 or any(v is None for v in values):
                continue
            date, time_str, name, status = values
            msg = f"🕒 {date} {time_str}\n👤 {name}\n📌 狀態：{status}"
            messages.append(msg)
        return "📋 最近出入紀錄：\n\n" + "\n\n".join(messages) if messages else "⚠️ 找不到有效資料。"
    except Exception as e:
        return f"⚠️ 讀取 Excel 發生錯誤：{e}"

# Webhook 入口，處理收到訊息
@app.route("/webhook", methods=["POST"])
def webhook():
    body = request.json
    events = body.get("events", [])
    for event in events:
        if event["type"] == "message":
            user_id = event["source"]["userId"]
            text = event["message"]["text"]
            reply_token = event["replyToken"]

            save_user_id(user_id)

            if text == "選單":
                reply_quick_menu(reply_token)
            elif text == "查詢最近紀錄":
                msg = get_latest_entries(5)
                reply_message(reply_token, msg)
            else:
                reply_message(reply_token, f"👋 收到訊息：「{text}」\n請輸入「選單」查看操作項目。")

    return "OK"

# 監控 Excel 新增資料並推播給所有用戶（文字＋圖片）
def monitor_excel(interval=10):
    global last_row_num
    last_row_num = 1
    if os.path.exists(STATE_FILE):
        with open(STATE_FILE, "r", encoding="utf-8") as f:
            try:
                last_row_num = int(f.read())
            except:
                last_row_num = 1

    while True:
        try:
            if not os.path.exists(EXCEL_FILE):
                print(f"❌ 找不到 {EXCEL_FILE}")
                time.sleep(interval)
                continue

            wb = load_workbook(EXCEL_FILE)
            sheet = wb.active
            current_last_row = sheet.max_row

            if current_last_row > last_row_num:
                print(f"🆕 偵測到新增資料：{current_last_row - last_row_num} 列")
                for row in range(last_row_num + 1, current_last_row + 1):
                    values = [cell.value for cell in sheet[row][:4]]
                    if len(values) < 4 or any(v is None for v in values):
                        continue
                    date, time_str, name, status = values
                    msg = f"📋 新出入紀錄\n🗓 {date} {time_str}\n👤 {name}\n📌 狀態：{status}"
                    push_to_all_users(msg, IMAGE_URL)
                last_row_num = current_last_row
                with open(STATE_FILE, "w", encoding="utf-8") as f:
                    f.write(str(last_row_num))
        except Exception as e:
            print("❌ 監控錯誤：", e)
        time.sleep(interval)

# 主程式入口
if __name__ == "__main__":
    port = 5000
    public_url = ngrok.connect(port)
    print(f"🌐 Ngrok 公開網址：{public_url}")
    print(f"📡 請設定 LINE Webhook 為：{public_url}/webhook")

    threading.Thread(target=monitor_excel, daemon=True).start()
    app.run(port=port, threaded=True)

